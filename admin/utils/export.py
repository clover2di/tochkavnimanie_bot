"""
Export utilities for generating Excel reports.
"""
import io
from datetime import datetime, timedelta
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def export_applications_to_xlsx(applications: List, nominations: dict = None) -> io.BytesIO:
    """
    Export applications to Excel file.
    
    Args:
        applications: List of Application objects with related User and Nomination
        nominations: Optional dict of nomination_id -> nomination_name for headers
    
    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "ID",
        "Дата подачи",
        "ФИО",
        "Telegram",
        "Город",
        "Школа",
        "Класс",
        "Этап",
        "Кол-во файлов",
        "Комментарий",
        "Голосовое"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for row_num, app in enumerate(applications, 2):
        # Parse files count
        files_count = 0
        if app.photos:
            try:
                import json
                photos = json.loads(app.photos)
                files_count += len(photos) if isinstance(photos, list) else 0
            except:
                pass
        if hasattr(app, 'files') and app.files:
            try:
                import json
                files = json.loads(app.files)
                files_count += len(files) if isinstance(files, list) else 0
            except:
                pass
        
        # Format datetime with +5 hours
        created_at = app.created_at
        if created_at:
            created_at = created_at + timedelta(hours=5)
            created_at_str = created_at.strftime("%d.%m.%Y %H:%M")
        else:
            created_at_str = ""
        
        row_data = [
            app.id,
            created_at_str,
            app.user.full_name or "",
            f"@{app.user.username}" if app.user.username else str(app.user.telegram_id),
            app.user.city or "",
            app.user.school or "",
            app.user.grade or "",
            app.nomination.name if app.nomination else "",
            files_count,
            app.comment_text or "",
            "Да" if app.voice_file_id else "Нет"
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    
    # Adjust column widths
    column_widths = [6, 16, 25, 18, 15, 30, 8, 20, 12, 40, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def export_participants_to_xlsx(participants: List) -> io.BytesIO:
    """
    Export participants to Excel file.
    
    Args:
        participants: List of dicts with 'user', 'application_count', 'last_application_date'
    
    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Участники"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "ID",
        "Telegram",
        "ФИО",
        "Город",
        "Школа",
        "Класс",
        "Кол-во заявок",
        "Последняя заявка",
        "Дата регистрации"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for row_num, item in enumerate(participants, 2):
        user = item['user']
        
        # Format dates with +5 hours
        last_app = item.get('last_application_date')
        if last_app:
            last_app = last_app + timedelta(hours=5)
            last_app_str = last_app.strftime("%d.%m.%Y %H:%M")
        else:
            last_app_str = ""
        
        created_at = user.created_at
        if created_at:
            created_at = created_at + timedelta(hours=5)
            created_at_str = created_at.strftime("%d.%m.%Y %H:%M")
        else:
            created_at_str = ""
        
        row_data = [
            user.id,
            f"@{user.username}" if user.username else str(user.telegram_id),
            user.full_name or "",
            user.city or "",
            user.school or "",
            user.grade or "",
            item.get('application_count', 0),
            last_app_str,
            created_at_str
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    
    # Adjust column widths
    column_widths = [6, 18, 25, 15, 30, 8, 12, 16, 16]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
